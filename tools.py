from os import path as osp
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from torchvision import transforms
from collections import Counter

import matplotlib.pyplot as plt
import numpy as np
import torch.nn as nn
import torch
import data_list
from torch.utils.data import DataLoader, random_split
from sklearn.metrics import confusion_matrix
import network
import mix
import loss
import model
import torch.nn.functional as F

class InvalidDataError(Exception):
    def __init__(self, variable_name):
        self.variable_name = variable_name
        super().__init__(f"变量{variable_name}的内容不合法")


# class Normalize1D:
#     def __init__(self, data):
#         self.mean = np.mean(data)
#         self.std = np.std(data)
#
#     def __call__(self, data):
#         """
#         对一维数据进行标准化操作
#
#         Parameters:
#         - data: 一维数据
#         """
#         normalized_data = (data - self.mean) / self.std
#         return normalized_data


class Normalize:
    def __init__(self, data=None, normalized_choose=None):
        self.mean = None if data is None else np.mean(data)
        self.std = None if data is None else np.std(data)
        self.normalized_choose = normalized_choose

    def __call__(self, data):
        """
        对数据进行标准化操作
        Parameters:
        - data: 数据，形状可以是 (N, H, W) 或 (H, W)
        """
        if data.ndim == 3:
            # 对每个的样本分别进行归一/标准化
            normalized_data = np.array([self._normalize(data[i]) for i in range(data.shape[0])])
        elif 1 <= data.ndim <= 2:
            # 如果数据是二维的，则直接归一/标准化
            normalized_data = self._normalize(data)
        else:
            raise ValueError("Input data must be 2D or 3D array.")
        return normalized_data

    def _normalize(self, data):
        """
        对单个样本进行标准化操作
        Parameters:
        - data: 单个样本数据，形状可以是 (H, W)
        """
        if self.normalized_choose == 'min_max':
            # 归一化
            normalized_data = (data - np.min(data)) / (np.max(data) - np.min(data))
        elif self.normalized_choose == 'standard':
            # 标准化
            normalized_data = (data - np.mean(data)) / np.std(data)
        elif self.normalized_choose == 'z_score':
            # 标准化
            normalized_data = (data - self.mean) / self.std
        elif self.normalized_choose == 'none':
            # 什么都不做
            normalized_data = data
        else:
            raise ValueError("Invalid normalized_choose value. Choose from 'min_max', 'standard', or 'none'.")
        return normalized_data


class Tensor1D:
    def __init__(self):
        self.type = torch.float

    def __call__(self, data):
        return torch.tensor(data, dtype=self.type)


class AddNoise:
    def __init__(self, mean=0, std=0.1):
        self.mean = mean
        self.std = std

    def __call__(self, data):
        noise = np.random.normal(self.mean, self.std, size=len(data))
        noisy_data = data + noise
        return noisy_data


class DataPreprocessor:
    """
    任务：数据预处理器，用于对数据进行标准化、添加噪声和转换为张量。

    参数：
    - data: 待处理的数据。
    - std: 标准差，用于添加噪声的标准差，默认为 0.1。
    - ctype: 数据类型，'TXT' 表示文本数据，默认为 'TXT'。
    - choose: 选择处理方式，'train' 表示训练模式，'test' 表示测试模式，默认为 'train'。

    初始化方法会根据数据类型和处理方式选择合适的处理器。处理器包括标准化器、噪声添加器和张量转换器。

    在训练模式下，数据将依次经过标准化、添加噪声和转换为张量；在测试模式下，数据只经过标准化和转换为张量。

    注意：目前仅支持 'TXT' 类型的数据预处理，其他类型的数据将会引发 InvalidDataError 异常。
    """

    def __init__(self, data=None, normalized_choose=None, std=0.1, ctype='TXT', choose='train'):
        """
        初始化方法，根据数据类型和处理方式选择合适的处理器。
        参数：
        - data: 待处理的数据。
        - std: 标准差，用于添加噪声的标准差，默认为 0.1。
        - ctype: 数据类型，'TXT' 表示文本数据，默认为 'TXT'。
        - choose: 选择处理方式，'train' 表示训练模式，'test' 表示测试模式，默认为 'train'。
        根据数据类型和处理方式选择合适的处理器。处理器包括标准化器、噪声添加器和张量转换器。
        """
        if ctype == 'TXT':
            self.normalize = Normalize(data, normalized_choose)
            self.to_tensor = Tensor1D()
            self.noise = AddNoise(std=std)
        else:
            raise InvalidDataError("type")

        self.choose = choose

    def __call__(self, data):
        """
        调用方法，根据选择的处理方式对数据进行预处理。
        参数：
        - data: 待处理的数据。
        返回值：
        - transformed_data: 经过预处理后的数据。
        在训练模式下，数据将依次经过标准化、添加噪声和转换为张量；在测试模式下，数据只经过标准化和转换为张量。
        """
        if self.choose == 'train':
            transforming = transforms.Compose([
                self.normalize,
                self.noise,
                self.to_tensor
            ])
            return transforming(data)
        else:
            transforming = transforms.Compose([
                self.normalize,
                self.to_tensor
            ])
            return transforming(data)


class find_position():
    def __init__(self, n):
        self.n = n
        self.matrix = [[-1] * n for _ in range(n)]
        count = 0
        for i in range(n * n):
            row, col = divmod(i, n)
            if row != col:
                self.matrix[row][col] = count
                count += 1

    def __call__(self, i, j):
        if i >= self.n or j >= self.n:
            raise ValueError("i或j超出定义")
        else:
            return self.matrix[i][j]


def Entropy(input_):
    """
    任务：计算输入张量的熵（entropy）。

    参数：
    - input_: 输入张量，形状为 [batch_size, num_classes]。

    返回值：
    - entropy: 每个样本的熵，形状为 [batch_size]。

    计算输入张量的熵，用于衡量每个样本的不确定性。这个函数首先引入一个小的正数 epsilon，以防止在计算对数时出现零值，
    从而避免产生无效的结果。然后按照熵的公式 H(x) = - Σ p(x) * log(p(x)) 对输入张量进行计算，并在横向（dim=1）对每个类别的熵进行求和，
    得到每个样本的整体熵。

    注意：输入张量的值应该在 [0, 1] 范围内，并且每个样本的概率和应该为 1。
    """
    bs = input_.size(0)
    # 引入一个小的正数 epsilon，用于防止在计算对数时出现零值，以避免产生无效的结果。
    epsilon = 1e-5
    # 这是计算熵的常见方式，即 H(x) = - Σ p(x) * log(p(x))。
    entropy = -input_ * torch.log(input_ + epsilon)
    # 对每个样本计算其横向（dim=1）的熵，即对每个类别的熵进行求和。这可以用来得到每个样本的整体熵。
    entropy = torch.sum(entropy, dim=1)
    return entropy


def op_copy(optimizer):
    """
    任务：复制优化器中的当前学习率到一个新的键 'lr0' 中。

    参数：
    - optimizer: 要复制学习率的优化器。

    返回值：
    - optimizer: 复制了学习率的优化器。

    复制优化器中的当前学习率到一个新的键 'lr0' 中。这样，每个参数组就保留了初始的学习率，以便之后进行学习率衰减时进行参考。
    """
    for param_group in optimizer.param_groups:
        # 将当前的学习率 'lr' 复制到一个初始学习率键 'lr0' 中。这样，每个参数组就保留了初始的学习率，以便之后进行学习率衰减时进行参考。
        param_group['lr0'] = param_group['lr']
    return optimizer


def lr_scheduler(optimizer, iter_num, max_iter, gamma=10, power=0.75):
    """
    任务：根据迭代次数动态调整优化器的学习率。

    参数：
    - optimizer: 要调整学习率的优化器。
    - iter_num: 当前迭代次数。
    - max_iter: 总的迭代次数。
    - gamma: 学习率衰减因子，默认为10。
    - power: 学习率衰减的幂次，默认为0.75。

    返回值：
    - 调整后的优化器。

    根据迭代次数动态调整优化器的学习率，使用指数衰减的方式。学习率随着迭代次数的增加而减小，以防止模型过拟合，并提高收敛速度。
    还会对优化器的 weight_decay、momentum 和 nesterov 参数进行设置，以进一步优化模型的训练效果。

    注意：调用该函数后需要重新计算梯度以应用新的学习率。
    """

    decay = (1 + gamma * iter_num / max_iter) ** (-power)  # 衰减系数
    for param_group in optimizer.param_groups:  # param_group['lr0']是初始学习率，通过decay进行衰减
        param_group['lr'] = param_group['lr0'] * decay
        # 权重衰减（Weight Decay）是一种正则化技术，旨在防止模型过拟合。它通过在损失函数中添加一个惩罚项，惩罚较大的权重值。
        param_group['weight_decay'] = 1e-3
        param_group['momentum'] = 0.9
        # Nesterov 动量是动量的一种变体，它在计算梯度之前先进行一次“预期”更新，然后在该位置计算梯度。
        param_group['nesterov'] = True
    return optimizer


def print_args(args):
    s = "==========================================\n"
    for arg, content in args.__dict__.items():
        s += "{}:{}\n".format(arg, content)
    return s


def print_excel(args, str_list, acc_list):
    """
    任务：根据给定的参数、字符串列表和正确率列表，生成一个包含这些信息的字典，用于写出对应的 Excel 表格。

    参数：
    - args: 包含程序运行参数的命名空间对象。
    - str_list: 包含字符串的列表，用于描述各项指标或参数。
    - acc_list: 包含正确率或其他指标值的列表。

    返回值：
    - excel_dict: 一个包含了时间、数据集、参数和正确率等信息的字典，用于生成 Excel 表格。
    """
    # 创建完整字典
    # 加上时间和数据集
    excel_dict = {'时间': args.nowtime, '数据集': args.data_name}
    # 加上参数
    excel_dict.update({string: '' for string in args.choose_dict})
    # 加上正确率
    excel_dict.update({string: '' for string in str_list})

    # 开始填充内容
    # 填充时间
    excel_dict["时间"] = args.nowtime

    # 填充参数
    for arg, content in args.__dict__.items():
        if arg in args.choose_dict:
            excel_dict[arg] = content
    pass

    # 填充正确率
    for index, key in enumerate(str_list):
        if acc_list[index] != -1:
            excel_dict[key] = round(acc_list[index], 2)

    return excel_dict


def write_to_excel(data_dict, excel_file):
    """
    任务：将给定的数据字典写入到指定的 Excel 文件中。

    参数：
    - data_dict: 包含数据的字典，其中键对应 Excel 表头，值对应要写入的数据。
    - excel_file: 要写入数据的 Excel 文件的路径。

    如果指定的 Excel 文件不存在，则创建一个新的 Excel 文件，并将数据写入其中；如果文件已存在，则加载现有的文件并将数据附加到最后一行。

    注意：数据字典中的键必须与 Excel 表格的表头相匹配，否则会导致数据写入错误的列。
    """
    if not osp.exists(excel_file):
        # 如果 Excel 文件不存在，创建一个新的 Excel 文件，并写入表头
        wb = Workbook()
        ws = wb.active
        ws.append(list(data_dict.keys()))  # 写入表头
    else:
        # 如果 Excel 文件已存在，加载现有的 Excel 文件
        wb = load_workbook(excel_file)
        ws = wb.active

    # 写入数据
    data_row = []
    for header in ws[1]:
        value = data_dict.get(header.value, '')  # 获取字典中对应键的值，如果键不存在，写入空值
        # 判断 value 的类型
        if isinstance(value, (int, float)):  # 如果是数值类型
            data_row.append(value)
        else:  # 如果是非数值类型
            data_row.append(str(value))
    ws.append(data_row)

    # 设置单元格样式
    for cell in ws[ws.max_row]:
        cell.font = Font(name='Times New Roman', size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 保存 Excel 文件
    wb.save(excel_file)


# 输出正确率矩阵
def print_acc(acc_matrix):
    correct_rate = [0] * len(acc_matrix)
    log_str = ""
    for z in range(len(acc_matrix)):
        total = acc_matrix[z].sum()  # 计算该行的总和
        if total != 0:
            correct_rate[z] = round(acc_matrix[z, z] / total * 100)
        else:
            correct_rate[z] = 0  # 当该行全为 0 时，准确率设为 0
            log_str += f"类型{z + 1} 没有样本或未被分类。\n"
            continue  # 跳过该类别的预测分布计算

        correct_rate_str = f"类型{z + 1} 正确率：{correct_rate[z]: 7.2f}%"

        # 计算预测分布，处理总和不为 0 的情况
        prediction_ratios = [f"{j + 1} {round(acc_matrix[z, j] / total * 100, 2): 7.2f}%" for j in range(len(acc_matrix))]
        prediction_ratios_str = ", ".join(prediction_ratios)

        log_str += f"{correct_rate_str} -> {prediction_ratios_str}\n"

    return correct_rate, log_str


def args_data_load(args, dataset_info, separate=True):
    # Prepare data
    train_batch_size = args.batch_size
    txt_lines = open(args.dset_path).readlines()

    # Determine dataset size
    total_size = len(txt_lines) if args.num == -1 else args.num * args.class_num

    # Initialize variables for dataset splits
    remaining_txt, _ = random_split(txt_lines, [total_size, len(txt_lines) - total_size])
    remaining_size = total_size
    data_loaders = {}

    # Split the dataset based on cut_ratios
    for name, ratio in dataset_info.items():
        if separate:
            split_size = int(ratio * total_size)
            remaining_size -= split_size
            split_txt, remaining_txt = random_split(remaining_txt, [split_size, remaining_size])
        else:
            ratio = ratio if name != "train" else 1
            split_size = int(ratio * total_size)
            remaining_size -= split_size
            split_txt, _ = random_split(remaining_txt, [split_size, remaining_size])
            remaining_size += split_size

        data_array = np.array(
            [item for path, _ in data_list.make_dataset(split_txt, None) for item in data_list.txt_loader(path)])

        dataset = data_list.DataList(
            split_txt,
            transform=DataPreprocessor(data=data_array, normalized_choose="z_score", std=0, choose=name),
            mode='TXT'
        )
        data_loaders[name] = DataLoader(
            dataset, batch_size=train_batch_size, shuffle=True, num_workers=args.worker, drop_last=False
        )

    return data_loaders


def cal_acc_matrix(labels, predictions):
    conf_matrix = confusion_matrix(labels, predictions)
    per_class_acc = conf_matrix.diagonal() / conf_matrix.sum(axis=1) * 100
    overall_acc = round(per_class_acc.mean(), 2)

    return overall_acc, conf_matrix


def cal_acc(loader, networks):
    # 用于在循环中控制是否是第一次测试。
    start_test = True

    with torch.no_grad():  # 在测试阶段，关闭梯度计算
        for data in loader:  # 遍历数据加载器中的每个批次数据
            inputs = data[0]  # 获取输入数据
            labels = data[1]  # 获取标签
            inputs, labels = inputs.cuda(), labels.cuda()  # 将输入数据移到GPU上（如果可用）

            outputs = inputs
            for network in networks:
                outputs = network(outputs)

            if start_test:
                all_output = outputs.float().cpu()
                all_label = labels.float().cpu()
                start_test = False
            else:
                all_output = torch.cat((all_output, outputs.float().cpu()), 0)
                all_label = torch.cat((all_label, labels.float().cpu()), 0)

    all_output = nn.Softmax(dim=1)(all_output)
    _, predictions = torch.max(all_output, 1)
    mean_entropy = torch.mean(Entropy(all_output)).item()
    overall_acc, conf_matrix = cal_acc_matrix(all_label.numpy(), predictions.numpy())

    return overall_acc, conf_matrix, mean_entropy


def read_model(args, model_path=None):

    # # set base network
    # netF = network.ResBase(res_name=args.net).cuda()
    netF = model.ResNet18_1D(in_channels=1).cuda()
    # netF = network.TransformerEncoder(input_dim=args.input_dim, window=args.window, stride=args.stride,
    #                                   dropout=args.dropout, n_head=args.n_head, n_layers=args.n_layer,
    #                                   p_choose=args.PEmbedding, embed=args.PTEmbedding).cuda()

    if args.layer != "None":
        netB = network.feat_bootleneck(type=args.layer, feature_dim=netF.in_features,
                                       bottleneck_dim=args.bottleneck).cuda()
        netC = network.feat_classifier(type=args.classifier, class_num=args.class_num,
                                       bottleneck_dim=args.bottleneck).cuda()
        networks = [netF, netB, netC]
    else:
        netC = network.feat_classifier(type=args.classifier, class_num=args.class_num,
                                       bottleneck_dim=netF.in_features).cuda()
        networks = [netF, netC]

    if model_path is not None:
        for idx, net in enumerate(networks[:-1], start=1):
            net.load_state_dict(torch.load(f"{model_path}/source_{idx}.pt"))
        networks[-1].load_state_dict(torch.load(f"{model_path}/source_C.pt"))

    for net in networks:
        net.eval()

    return networks


class mix_train():
    def __init__(self):
        self.mix_iter_num = 0

    def __call__(self, args, networks, inputs, labels):
        mix_classifier_losss = []

        # 分类损失计算
        args_cls_par = getattr(args, 'cls_par', 1)

        if args_cls_par > 0:
            for i in range(int(args.mix) + 1):
                if np.random.rand(1) < args.mix - i:  # 进行 CutMix 的概率为 args.mix
                    self.mix_iter_num += 1
                    inputs_mix, pred_a, pred_b, lam = mix.cutmix_1d(inputs, labels)

                    # 前向传播
                    outputs_mix = inputs_mix
                    for network in networks:
                        outputs_mix = network(outputs_mix)

                    criterion = loss.CrossEntropyLabelSmooth(
                        num_classes=args.class_num, epsilon=args.smooth, reduction=False)
                    loss_value = mix.cutmix_criterion(criterion, outputs_mix, pred_a, pred_b, lam)
                    mix_classifier_losss.append(loss_value * args_cls_par)

        if mix_classifier_losss:
            mix_classifier_losss = torch.stack(mix_classifier_losss).view(-1)
        else:
            mix_classifier_losss = torch.tensor([0.0]).cuda()

        return mix_classifier_losss


def process_inputs(inputs, process_type):
    if process_type == "reverse":
        # 翻转数据
        inputs = torch.flip(inputs, dims=[-1])
    elif process_type == "mask":
        # 随机遮盖1/10的样本
        length = inputs.shape[-1]
        mask_size = length // 10  # 需要遮盖的总元素数
        # 随机选择需要遮盖的索引
        mask_indices = np.random.choice(length, mask_size, replace=False)  # 随机选出不重复的索引
        # 将这些索引位置的元素置为0
        inputs[..., mask_indices] = 0
    elif process_type == "shift":
        # 随机位移样本前后1/5
        length = inputs.shape[-1]
        shift_size = length // 5
        shift_amount = -shift_size if np.random.choice([True, False]) else shift_size
        inputs = torch.roll(inputs, shifts=shift_amount, dims=-1)
    # 如果为 none 或者其他值则不处理
    return inputs


def check_columns(arr):
    # 获取二维数组的行数和列数
    num_rows = arr.shape[0]
    num_cols = arr.shape[1]

    # 如果行数为1，直接返回该行作为结果
    if num_rows == 1:
        return arr.flatten()

    # 计算行数的一半向上取整
    threshold = num_rows // 2
    # 初始化一个长度为列数的一维数组
    result = np.empty(num_cols, dtype=int)
    # 遍历每一列
    for i in range(num_cols):
        # 统计每列中元素的出现次数
        column = arr[:, i]
        counts = Counter(column)
        # 检查是否有数字出现次数超过行数的一半
        found = False
        for num, count in counts.items():
            if count > threshold:
                result[i] = num
                found = True
                break
        # 如果没有数字的出现次数超过行数的一半，则设置为-1
        if not found:
            result[i] = -1
    return result


def obtain_bank(loader, networks, args, process_type="none"):
    num_sample = len(loader.dataset)

    fea_bank = torch.randn(num_sample, args.bottleneck if args.layer != "None" else args.window)
    label_bank = torch.randn(num_sample, dtype=torch.float32)
    score_bank = torch.randn(num_sample, args.class_num).cuda()

    # 该部分不需要反向传播
    with torch.no_grad():
        for data in iter(loader):
            inputs = data[0]
            labels = data[1]
            indx = data[-1]
            inputs = inputs.cuda()

            # 对inputs进行处理
            inputs = process_inputs(inputs, process_type)  # 根据参数进行处理

            # 将输入数据传递到 netF 进行特征提取，然后通过 netB 得到特征
            features = inputs
            for net in networks[:-1]:
                features = net(features)

            # 对特征进行归一化
            output_norm = F.normalize(features)
            # 使用 Softmax 得到分类概率
            outputs = nn.Softmax(-1)(networks[-1](features))

            # 将这些特征存储在 fea_bank 中
            fea_bank[indx] = output_norm.detach().clone().cpu()
            # 将真标签存放在 label_bank 中
            label_bank[indx] = labels.float()
            # 将分类概率存储在 score_bank 中
            score_bank[indx] = outputs.detach().clone()  # .cpu()

    # print("仓库数据更新完成")

    return fea_bank, label_bank, score_bank

